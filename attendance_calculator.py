import streamlit as st
import pandas as pd
from datetime import datetime
import re
import io
import os

# Page Configuration
st.set_page_config(
    page_title="Office Attendance Calculator",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# Custom CSS
st.markdown("""
<style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
        max-width: 1200px !important;
        margin: 0 auto !important;
    }
    
    .app-title {
        font-size: 2rem;
        font-weight: 700;
        color: #4A90E2;
        margin-bottom: 0.3rem;
    }
    .app-subtitle {
        font-size: 0.95rem;
        color: #888;
        margin-bottom: 1.5rem;
        padding-bottom: 1rem;
        border-bottom: 2px solid #4A90E2;
    }
    
    .section-title {
        font-size: 1.2rem;
        font-weight: 600;
        margin: 1.2rem 0 0.8rem 0;
    }
    
    .upload-title {
        font-size: 1.1rem;
        font-weight: 600;
        margin-bottom: 0.5rem;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
        border-bottom: 2px solid #4A90E2;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: transparent;
        border-radius: 6px 6px 0 0;
        padding: 10px 24px;
        font-weight: 600;
        font-size: 0.95rem;
    }
    .stTabs [aria-selected="true"] {
        background-color: #4A90E2;
        color: white !important;
    }
    
    .success-box {
        background-color: rgba(40, 167, 69, 0.15);
        border-left: 4px solid #28a745;
        padding: 12px 16px;
        border-radius: 4px;
        margin: 12px 0;
    }
    
    .info-box {
        background-color: rgba(74, 144, 226, 0.12);
        border-left: 4px solid #4A90E2;
        padding: 14px 18px;
        border-radius: 4px;
        margin: 12px 0;
    }
    
    [data-testid="stMetric"] {
        background-color: rgba(74, 144, 226, 0.08);
        padding: 15px;
        border-radius: 8px;
        border: 1px solid rgba(74, 144, 226, 0.2);
    }
    
    .stDownloadButton button {
        background-color: #4A90E2;
        color: white;
        border: none;
        font-weight: 600;
    }
    .stDownloadButton button:hover {
        background-color: #357ABD;
        color: white;
    }
    
    .legend-box {
        background-color: rgba(74, 144, 226, 0.08);
        border: 1px solid rgba(74, 144, 226, 0.3);
        padding: 12px 16px;
        border-radius: 6px;
        margin: 10px 0;
        font-size: 0.9rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="app-title">Office Attendance Calculator</div>', unsafe_allow_html=True)
st.markdown('<div class="app-subtitle">Upload monthly attendance sheet to generate detailed working hours report</div>', unsafe_allow_html=True)


def parse_time(time_str):
    try:
        return datetime.strptime(str(time_str).strip(), "%H:%M")
    except:
        return None


def calculate_day_hours(times):
    if not times or len(times) < 2:
        if len(times) == 1:
            return 0.0, "Single Punch"
        return 0.0, "No Data"

    parsed = [parse_time(t) for t in times]
    parsed = [p for p in parsed if p is not None]

    if len(parsed) < 2:
        return 0.0, "Invalid"

    first_in = parsed[0]
    last_out = parsed[-1]
    total_seconds = (last_out - first_in).total_seconds()

    if len(parsed) >= 4 and len(parsed) % 2 == 0:
        working_seconds = 0
        for i in range(0, len(parsed), 2):
            if i + 1 < len(parsed):
                diff = (parsed[i + 1] - parsed[i]).total_seconds()
                if diff > 0:
                    working_seconds += diff
        total_seconds = working_seconds

    if total_seconds < 0:
        return 0.0, "Invalid"

    hours = total_seconds / 3600
    return round(hours, 2), "Complete"


def format_hours_to_hm(hours):
    """Convert decimal hours to 'Hh MMm' format"""
    if hours == 0:
        return "0h 00m"
    
    negative = hours < 0
    hours = abs(hours)
    
    h = int(hours)
    m = int(round((hours - h) * 60))
    if m == 60:
        h += 1
        m = 0
    
    result = f"{h}h {m:02d}m"
    return f"-{result}" if negative else result


def read_excel_file(file_path, file_ext):
    all_sheets_data = {}

    if file_ext.lower() == '.xls':
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet_name in wb.sheet_names():
            sheet = wb.sheet_by_name(sheet_name)
            rows = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    val = cell.value
                    if cell.ctype == 3:
                        try:
                            time_tuple = xlrd.xldate_as_tuple(val, wb.datemode)
                            val = f"{time_tuple[3]:02d}:{time_tuple[4]:02d}"
                        except:
                            val = str(val)
                    row.append(val if val != '' else None)
                rows.append(tuple(row))
            all_sheets_data[sheet_name] = rows
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            all_sheets_data[sheet_name] = rows

    return all_sheets_data


def extract_employee_data(file_path, file_ext):
    all_sheets_data = read_excel_file(file_path, file_ext)
    all_employees = []
    seen_employees = set()

    for sheet_name, rows in all_sheets_data.items():
        if sheet_name.lower() == 'summary':
            continue

        i = 0
        while i < len(rows):
            row = rows[i]
            if not row:
                i += 1
                continue

            row_str = ' '.join([str(c) if c is not None else '' for c in row])

            if 'No' in row_str and 'Name' in row_str and 'Dept' in row_str:
                emp_no = ""
                emp_name = ""

                for idx, cell in enumerate(row):
                    if cell and 'No' in str(cell) and ':' in str(cell):
                        for j in range(idx + 1, len(row)):
                            if row[j] is not None and str(row[j]).strip():
                                emp_no = str(row[j]).strip()
                                break
                    if cell and 'Name' in str(cell) and ':' in str(cell):
                        for j in range(idx + 1, len(row)):
                            if row[j] is not None and str(row[j]).strip():
                                emp_name = str(row[j]).strip()
                                break

                emp_key = f"No.{emp_no}-{emp_name}"

                if emp_key in seen_employees:
                    i += 1
                    continue
                seen_employees.add(emp_key)

                current_days_data = {}

                if i + 1 < len(rows):
                    time_row = rows[i + 1]
                    for day_idx, cell in enumerate(time_row[:31], start=1):
                        if cell is not None:
                            cell_str = str(cell).strip()
                            times = re.findall(r'\d{1,2}:\d{2}', cell_str)
                            if times:
                                current_days_data[day_idx] = times

                all_employees.append({
                    'emp_no': emp_no,
                    'emp_name': emp_name.title(),
                    'days': current_days_data
                })
                i += 2
                continue

            i += 1

    return all_employees


def generate_report(employees):
    summary_data = []
    detailed_data = []

    for emp in employees:
        emp_no = emp['emp_no']
        emp_name = emp['emp_name']
        days = emp['days']

        total_hours = 0
        present_days = 0

        for day_num in sorted(days.keys()):
            times = days[day_num]
            hours, status = calculate_day_hours(times)

            if hours > 0:
                present_days += 1
                total_hours += hours

            detailed_data.append({
                'Emp No': emp_no,
                'Employee Name': emp_name,
                'Day': day_num,
                'Punch Times': ', '.join(times),
                'Total Punches': len(times),
                'Working Time': format_hours_to_hm(hours),
                'Status': status,
                '_hours_raw': hours
            })

        avg_hours = round(total_hours / present_days, 2) if present_days > 0 else 0
        expected_hours = present_days * 8
        difference = round(total_hours - expected_hours, 2)

        # Status label
        if present_days == 0:
            status_label = "Absent"
        elif difference >= 0:
            status_label = "On Target"
        else:
            status_label = "Below Target"

        summary_data.append({
            'Emp No': emp_no,
            'Employee Name': emp_name,
            'Present Days': present_days,
            'Total Working Time': format_hours_to_hm(total_hours),
            'Avg Time/Day': format_hours_to_hm(avg_hours),
            'Expected Time': format_hours_to_hm(expected_hours),
            'Difference': format_hours_to_hm(difference),
            'Status': status_label,
            '_total_raw': total_hours,
            '_diff_raw': difference
        })

    summary_df = pd.DataFrame(summary_data)
    detailed_df = pd.DataFrame(detailed_data)
    return summary_df, detailed_df


# File Upload Section
st.markdown('<div class="upload-title">Upload Attendance File</div>', unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "Supported formats: .xls and .xlsx",
    type=['xlsx', 'xls'],
    help="Upload the monthly attendance log exported from your biometric system"
)

if uploaded_file:
    file_name = uploaded_file.name
    file_ext = os.path.splitext(file_name)[1].lower()

    with st.spinner('Processing attendance data, please wait...'):
        temp_file = f"temp_attendance{file_ext}"
        with open(temp_file, "wb") as f:
            f.write(uploaded_file.getbuffer())

        try:
            employees = extract_employee_data(temp_file, file_ext)

            if not employees:
                st.error("No employee data found. Please verify the file format.")
            else:
                st.markdown(
                    f'<div class="success-box"><b>Success:</b> Data extracted for <b>{len(employees)} employees</b> from <b>{file_name}</b></div>',
                    unsafe_allow_html=True
                )

                summary_df, detailed_df = generate_report(employees)

                tab1, tab2, tab3 = st.tabs([
                    "Summary Dashboard",
                    "Day-wise Details",
                    "Individual Report"
                ])

                # TAB 1: SUMMARY
                with tab1:
                    st.markdown('<div class="section-title">Employee Summary</div>', unsafe_allow_html=True)
                    
                    # Column explanation
                    st.markdown("""
                    <div class="legend-box">
                    <b>Column Guide:</b><br>
                    <b>Present Days</b> - Number of days employee came to office<br>
                    <b>Total Working Time</b> - Total time spent in office for the month<br>
                    <b>Avg Time/Day</b> - Average working time per day (Total / Present Days)<br>
                    <b>Expected Time</b> - Required time based on 8 hours per present day<br>
                    <b>Difference</b> - Extra or short time compared to expected (negative means less)<br>
                    <b>Status</b> - On Target (met 8hr requirement) or Below Target
                    </div>
                    """, unsafe_allow_html=True)

                    search = st.text_input(
                        "Search employee by name",
                        placeholder="Type to filter employees...",
                        label_visibility="collapsed"
                    )

                    display_df = summary_df.copy()
                    if search:
                        display_df = display_df[
                            display_df['Employee Name'].str.contains(search, case=False, na=False)
                        ]

                    # Drop internal columns for display
                    display_cols = ['Emp No', 'Employee Name', 'Present Days', 'Total Working Time',
                                    'Avg Time/Day', 'Expected Time', 'Difference', 'Status']

                    st.dataframe(
                        display_df[display_cols],
                        use_container_width=True,
                        height=500,
                        hide_index=True,
                        column_config={
                            "Emp No": st.column_config.TextColumn("Emp No", width="small"),
                            "Employee Name": st.column_config.TextColumn("Employee Name", width="medium"),
                            "Present Days": st.column_config.NumberColumn("Present Days", width="small"),
                            "Total Working Time": st.column_config.TextColumn("Total Working Time"),
                            "Avg Time/Day": st.column_config.TextColumn("Avg Time/Day"),
                            "Expected Time": st.column_config.TextColumn("Expected Time"),
                            "Difference": st.column_config.TextColumn("Difference"),
                            "Status": st.column_config.TextColumn("Status"),
                        }
                    )

                    st.markdown('<div class="section-title">Download Reports</div>', unsafe_allow_html=True)

                    export_summary = summary_df[display_cols].copy()
                    export_detailed = detailed_df.drop(columns=['_hours_raw'], errors='ignore')

                    col_dl1, col_dl2 = st.columns(2)
                    with col_dl1:
                        csv_summary = export_summary.to_csv(index=False)
                        st.download_button(
                            "Download Summary (CSV)",
                            csv_summary,
                            "attendance_summary.csv",
                            "text/csv",
                            use_container_width=True
                        )
                    with col_dl2:
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            export_summary.to_excel(writer, sheet_name='Summary', index=False)
                            export_detailed.to_excel(writer, sheet_name='Day-wise Details', index=False)
                        st.download_button(
                            "Download Full Report (Excel)",
                            output.getvalue(),
                            "attendance_full_report.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                # TAB 2: DAY-WISE
                with tab2:
                    st.markdown('<div class="section-title">Day-wise Attendance Details</div>', unsafe_allow_html=True)

                    col_f1, col_f2 = st.columns([2, 1])
                    with col_f1:
                        emp_filter = st.multiselect(
                            "Filter by Employee",
                            options=summary_df['Employee Name'].tolist(),
                            placeholder="Select employees..."
                        )
                    with col_f2:
                        status_filter = st.multiselect(
                            "Filter by Status",
                            options=detailed_df['Status'].unique().tolist(),
                            placeholder="Select status..."
                        )

                    filtered_df = detailed_df.copy()
                    if emp_filter:
                        filtered_df = filtered_df[filtered_df['Employee Name'].isin(emp_filter)]
                    if status_filter:
                        filtered_df = filtered_df[filtered_df['Status'].isin(status_filter)]

                    st.markdown(f"**Showing {len(filtered_df)} records**")

                    display_detailed_cols = ['Emp No', 'Employee Name', 'Day', 'Punch Times', 
                                             'Total Punches', 'Working Time', 'Status']

                    st.dataframe(
                        filtered_df[display_detailed_cols],
                        use_container_width=True,
                        height=500,
                        hide_index=True
                    )

                    csv_detail = filtered_df[display_detailed_cols].to_csv(index=False)
                    st.download_button(
                        "Download Detailed Report (CSV)",
                        csv_detail,
                        "attendance_detailed.csv",
                        "text/csv"
                    )

                # TAB 3: INDIVIDUAL
                with tab3:
                    st.markdown('<div class="section-title">Individual Employee Report</div>', unsafe_allow_html=True)

                    emp_options = summary_df.apply(
                        lambda x: f"{x['Employee Name']} (No.{x['Emp No']})", axis=1
                    ).tolist()

                    selected = st.selectbox("Select Employee", emp_options)

                    if selected:
                        selected_no = selected.split("No.")[-1].replace(")", "").strip()
                        emp_summary = summary_df[summary_df['Emp No'] == selected_no].iloc[0]
                        emp_details = detailed_df[detailed_df['Emp No'] == selected_no]

                        st.markdown(f"#### {emp_summary['Employee Name']} - Employee No. {emp_summary['Emp No']}")

                        col1, col2, col3, col4 = st.columns(4)
                        col1.metric("Present Days", emp_summary['Present Days'])
                        col2.metric("Total Working Time", emp_summary['Total Working Time'])
                        col3.metric("Avg Time/Day", emp_summary['Avg Time/Day'])
                        col4.metric("Difference", emp_summary['Difference'])

                        st.markdown('<div class="section-title">Daily Attendance Log</div>', unsafe_allow_html=True)
                        
                        ind_cols = ['Day', 'Punch Times', 'Total Punches', 'Working Time', 'Status']
                        st.dataframe(
                            emp_details[ind_cols],
                            use_container_width=True,
                            hide_index=True,
                            height=450
                        )

                        csv_ind = emp_details[ind_cols].to_csv(index=False)
                        st.download_button(
                            f"Download {emp_summary['Employee Name']} Report",
                            csv_ind,
                            f"attendance_{emp_summary['Employee Name']}.csv",
                            "text/csv"
                        )

        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            with st.expander("Show error details"):
                st.exception(e)
        finally:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass

else:
    st.markdown("""
    <div class="info-box">
    <b>How to use this tool:</b><br><br>
    1. Export the monthly attendance log from your biometric system in .xls or .xlsx format<br>
    2. Click <b>Browse files</b> button above and select the file<br>
    3. View reports across three tabs: <b>Summary</b>, <b>Day-wise Details</b>, and <b>Individual Report</b><br>
    4. Download reports in CSV or Excel format for record keeping
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-title">Calculation Logic</div>', unsafe_allow_html=True)
    st.markdown("""
    - **2 punches per day:** First punch as IN time, last punch as OUT time
    - **4 or more punches:** System detects lunch breaks and excludes break time
    - **Single or odd punches:** Marked as incomplete, first and last punch considered
    - **Expected working hours:** 8 hours per present day (standard target)
    """)